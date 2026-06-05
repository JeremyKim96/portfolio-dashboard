"""portfolio.xlsx 템플릿 생성/보강.

- 파일 없으면: 4시트(coins, cash, manual, meta) 새로 생성
- 파일 있으면: 누락된 시트만 추가 (기존 데이터·시트 보존)

저장 위치: core.config.PORTFOLIO_XLSX (사용자 데스크탑의 자산 관리 폴더).
"""
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

from core.config import PORTFOLIO_XLSX

OUT = PORTFOLIO_XLSX
OUT.parent.mkdir(parents=True, exist_ok=True)

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="305496")


def _style_header(ws, headers):
    for c in ws[1]:
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
    for i, h in enumerate(headers, start=1):
        col_letter = ws.cell(row=1, column=i).column_letter
        ws.column_dimensions[col_letter].width = max(12, len(h) * 2 + 2)


def _add_sheet(wb, name, headers, rows):
    ws = wb.create_sheet(name)
    ws.append(headers)
    for r in rows:
        ws.append(r)
    _style_header(ws, headers)


SHEET_SPECS = {
    "coins": (
        ["심볼", "한글명", "수량", "평단가(KRW)"],
        [["KRW-BTC", "비트코인", 0.0, 0],
         ["KRW-ETH", "이더리움", 0.0, 0]],
    ),
    "cash": (
        ["구분", "금액", "메모"],
        [["주식 예수금", 0, "신한투자증권 (KRW 환산)"],
         ["코인 예수금", 0, "업비트 원화"],
         ["대출잔액",   0, "음수로! 예: -15000000"]],
    ),
    "manual": (
        ["분류", "종목명", "수량", "평단가", "통화", "현재가"],
        [],
    ),
    "family": (
        ["항목", "금액", "메모"],
        [["승환 보증금",  0, "전세/월세 보증금 등"],
         ["다은 보증금",  0, "전세/월세 보증금 등"],
         ["다은 투자자산", 0, "다은이 운용하는 투자자산 합계"],
         ["대출",        0, "대출/부채 — 금액은 양수로 입력하면 자동 차감됨"]],
    ),
    "meta": (
        ["티커", "한글명", "레버리지", "자산버킷", "목표비중(%)", "비고"],
        [["TQQQ",    "나스닥3X",      3.0, "빅테크",      10, ""],
         ["SOXL",    "반도체3X",      3.0, "AI/반도체",   15, ""],
         ["QLD",     "나스닥2X",      2.0, "빅테크",      10, ""],
         ["QQQ",     "나스닥100",     1.0, "빅테크",      20, ""],
         ["IGV",     "북미 소프트웨어",1.0, "빅테크",       5, ""],
         ["MAGS",    "매그니피센트7",  1.0, "빅테크",       5, ""],
         ["TSLA",    "테슬라",        1.0, "빅테크",       5, ""],
         ["MSFT",    "마이크로소프트", 1.0, "빅테크",       3, ""],
         ["GOOGL",   "알파벳A",       1.0, "빅테크",       3, ""],
         ["ADBE",    "어도비",        1.0, "빅테크",       2, ""],
         ["ORCL",    "오라클",        1.0, "빅테크",       2, ""],
         ["PLTR",    "팔란티어",      1.0, "AI/반도체",   2, ""],
         ["ETH",     "이더리움",      1.0, "암호화폐",    10, ""],
         ["BTC",     "비트코인",      1.0, "암호화폐",     5, ""]],
    ),
}


if not OUT.exists():
    wb = Workbook()
    wb.remove(wb.active)
    for name, (headers, rows) in SHEET_SPECS.items():
        _add_sheet(wb, name, headers, rows)
    wb.save(OUT)
    print(f"[ok] 새로 생성: {OUT}")
else:
    wb = load_workbook(OUT)
    added = []
    for name, (headers, rows) in SHEET_SPECS.items():
        if name not in wb.sheetnames:
            _add_sheet(wb, name, headers, rows)
            added.append(name)
    if added:
        wb.save(OUT)
        print(f"[ok] 시트 추가: {', '.join(added)} → {OUT}")
    else:
        print(f"[skip] 모든 시트 존재: {OUT}")
